#!/usr/bin/env python

from ..api import get_all_templates
import openpyxl as pyxl

if __name__ == '__main__':
    colors = [
        pyxl.styles.colors.Color('00FFFF'),  # 1. Aqua
        pyxl.styles.colors.Color('7FFFD4'),  # 2. Aquamarine
        pyxl.styles.colors.Color('DAA520'),  # 3. GoldenRod
        pyxl.styles.colors.Color('F5F5DC'),  # 4. Beige
        pyxl.styles.colors.Color('8FBC8F'),  # 5. DarkSeaGreen
        pyxl.styles.colors.Color('B0C4DE'),  # 6. LightSteelBlue
        pyxl.styles.colors.Color('A52A2A'),  # 7. Brown
        pyxl.styles.colors.Color('DEB887'),  # 8. BurlyWood
        pyxl.styles.colors.Color('5F9EA0'),  # 9. CadetBlue
        pyxl.styles.colors.Color('7FFF00'),  # 10. Chartreuse
        pyxl.styles.colors.Color('D2691E'),  # 11. Chocolate
        pyxl.styles.colors.Color('FFB6C1'),  # 12. LightPink
        pyxl.styles.colors.Color('6495ED'),  # 13. CornflowerBlue
        pyxl.styles.colors.Color('FFF8DC'),  # 14. Cornsilk
        pyxl.styles.colors.Color('FFA07A'),  # 15. LightSalmon
    ]
    cell_fills = []
    for color in colors:
        cell_fills.append(pyxl.styles.fills.PatternFill(patternType='solid', fgColor=color))
    fills_len = len(cell_fills)
    templates = get_all_templates()
    wb = pyxl.Workbook()
    ws = wb.active
    ws.title = 'Templates'
    ws.append(['Index', 'Template Name', 'Type', 'Description', 'Transforms Sample', 'Destructive', 'temp1'])
    current_fill_index = 0
    dv = pyxl.worksheet.datavalidation.DataValidation(type="list", formula1="temp1,temp2,temp3", allow_blank=True)
    dv.error = "You entry not in list"
    dv.errorTitle = "Invalid error"
    dv.prompt = "Please select from temperature list"
    dv.promptTitle = "Temperature Selection"
    ws.add_data_validation(dv)
    for index, template in enumerate(templates):
        template_ws = wb.create_sheet(template.name)
        transforms = 'No'
        if template.input_data['does_transform']:
            transforms = 'Yes'
        destructive = 'No'
        if template.input_data['destructive']:
            destructive = 'Yes'
        ws.append([index, template.name, template.input_data['process_type'],
                   template.description, transforms, destructive, "temp(1)"])
        template_ws.append(['PROC: ' + template.name])
        setup_params = template.input_data['setup'][0]['properties']
        param_headers = []
        params = []
        for p in setup_params:
            param_headers.append('PARAM')
            params.append(p['name'])
        template_ws.append(param_headers)
        template_ws.append(params)
        if current_fill_index == fills_len:
            current_fill_index = 0
        for column_cells in template_ws.columns:
            length = max(len(str(cell.value) or "") for cell in column_cells)
            template_ws.column_dimensions[column_cells[0].column].width = length + 3
            for cell in column_cells:
                cell.fill = cell_fills[current_fill_index]
        current_fill_index += 1
    for column_cells in ws.columns:
        length = max(len(str(cell.value) or "") for cell in column_cells)
        ws.column_dimensions[column_cells[0].column].width = length + 3
    last_column = ws['G']
    first_cell = last_column[0]
    dv.add(first_cell)
    wb.save(filename='/tmp/MC-Templates.xlsx')
