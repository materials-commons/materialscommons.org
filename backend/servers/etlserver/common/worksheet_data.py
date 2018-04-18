import os
import openpyxl


class ExcelIO:
    def __init__(self):
        self.workbook = None
        self.current_worksheet = None
        # used for testing
        self._force_project_name_on_read = None
        self._force_experiment_name_on_read = None

    def read_entire_data_from_current_sheet(self):
        sheet = self.current_worksheet
        data = []
        max_last_data_col = 0
        for row in sheet.iter_rows():
            empty_row = True
            values = []
            for cell in row:
                empty_row = empty_row and (not cell.value)
            if empty_row:
                print("encountered empty row at row_index = " + str(len(data)) + ".  " +
                      "Assuming end of data at this location")
                break
            for cell in row:
                value = cell.value
                try:
                    if str(value).strip() == "" or ("n/a" in str(value).strip()):
                        value = None
                except UnicodeEncodeError:
                    pass
                if value and len(values) >= max_last_data_col:
                    max_last_data_col = len(values) + 1
                values.append(value)
            data.append(values)
        if len(data[0]) > max_last_data_col:
            print("encountered empty col at col_index = " + str(max_last_data_col) + ".  " +
                  "Assuming end of data at this location")
            remake_data = []
            for row in range(0, len(data)):
                values = []
                data_row = data[row]
                for col in range(0, max_last_data_col):
                    value = data_row[col]
                    values.append(value)
                remake_data.append(values)
            data = remake_data
        if self._force_project_name_on_read:
            data[0][0] = "PROJ: " + self._force_project_name_on_read
        if self._force_experiment_name_on_read:
            data[1][0] = "EXP: " + self._force_experiment_name_on_read
        return data

    def write_data(self, path, data_row_list, worksheet_name=None):
        self.workbook = openpyxl.Workbook()
        self.current_worksheet = self.workbook.worksheets[0]
        if worksheet_name:
            self.current_worksheet.name = worksheet_name
        for row in range(0, len(data_row_list)):
            data_row = data_row_list[row]
            for col in range(0, len(data_row)):
                data_item = data_row[col]
                self.current_worksheet.cell(column=col + 1, row=row + 1, value=data_item)
        self.workbook.save(filename=path)

    def read_workbook(self, path):
        if not os.path.isfile(path):
            raise FileNotFoundError(path)
        self.workbook = openpyxl.load_workbook(filename=path)

    def sheet_name_list(self):
        return self.workbook.sheetnames

    def set_current_worksheet_by_index(self, index):
        sheet_name = self.workbook.sheetnames[index]
        self.current_worksheet = self.workbook[sheet_name]
        return self.current_worksheet

    def close(self):
        self.workbook.close()

    # these methods used for testing
    def force_project_name_for_testing(self, project_name):
        self._force_project_name_on_read = project_name

    # these methods used for testing
    def force_experiment_name_for_testing(self, experiment_name):
        self._force_experiment_name_on_read = experiment_name
